import openpyxl
from openpyxl import *
from tkinter import *
wb = load_workbook('C:\\Users\ADMIN\Desktop\\EXCEL.xlsx')
sheet = wb.active
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = "NAME"
    sheet.cell(row=1, column=2).value = "COURSE"
    sheet.cell(row=1, column=3).value = "SEMESTER"
    sheet.cell(row=1, column=4).value = "FORM NUMBER"
    sheet.cell(row=1, column=5).value = "CONTACT NUMBER"
    sheet.cell(row=1, column=6).value = "EMAIL ID"
    sheet.cell(row=1, column=7).value = "ADDRESS"
def focus1(event):
    COURSE_field.focus_set()
def focus2(event):
    SEMESTER_field.focus_set()
def focus3(event):
    FORM_NO_field.focus_set()
def focus4(event):
    CONTACT_NO_field.focus_set()
def focus5(event):
    EMAIL_ID_field.focus_set()
def focus6(event):
    ADDRESS_field.focus_set()
def clear():
    NAME_field.delete(0,END)
    COURSE_field.delete(0, END)
    SEMESTER_field.delete(0, END)
    FORM_NO_field.delete(0, END)
    CONTACT_NO_field.delete(0, END)
    EMAIL_ID_field.delete(0, END)
    ADDRESS_field.delete(0, END)
def insert():
    if(NAME_field.get() == "" and
        COURSE_field.get() == ""and
        SEMESTER_field.get() == ""and
        FORM_NO_field.get() == ""and
        CONTACT_NO_field.get == "" and
        EMAIL_ID_field.get() == ""and
        ADDRESS_field.get() == ""):
        print("EMPTY INPUT")
    else:
        current_row = sheet.max_row
        sheet.cell(row=current_row + 1, column=1).value = NAME_field.get()
        sheet.cell(row=current_row + 1, column=2).value = COURSE_field.get()
        sheet.cell(row=current_row + 1, column=3).value = SEMESTER_field.get()
        sheet.cell(row=current_row + 1, column=4).value = FORM_NO_field.get()
        sheet.cell(row=current_row + 1, column=5).value = CONTACT_NO_field.get()
        sheet.cell(row=current_row + 1, column=6).value = EMAIL_ID_field.get()
        sheet.cell(row=current_row + 1, column=7).value = ADDRESS_field.get()
        wb.save('C:\\Users\\ADMIN\\Desktop\\EXCEL.xlsx')
        NAME_field.focus_set()
        clear()
if __name__ == "__main__":
    root = Tk()
    root.configure(background = "yellow")
    root.title("REGISTRATION FORM")
    root.geometry("500x300")
    excel()
    HEADING = Label(root,text="REGISTRATION FORM",bg="yellow")
    NAME = Label(root, text="NAME",bg="yellow")
    COURSE = Label(root, text="COURSE",bg="yellow")
    SEMESTER = Label(root, text="SEMESTER",bg="yellow")
    FORM_NO = Label(root, text="FORM NO.",bg="yellow")
    CONTACT_NO = Label(root, text="CONTACT NO.",bg="yellow")
    EMAIL_ID = Label(root,text ="EMAIL ID",bg="yellow")
    ADDRESS = Label(root, text="ADDRESS",bg="yellow")

    HEADING.grid(row=0, column=1)
    NAME.grid(row=1, column=0)
    COURSE.grid(row=2, column=0)
    SEMESTER.grid(row=3, column=0)
    FORM_NO.grid(row=4, column=0)
    CONTACT_NO.grid(row=5, column=0)
    EMAIL_ID.grid(row=6, column=0)
    ADDRESS.grid(row=7, column=0)

    NAME_field = Entry(root)
    COURSE_field = Entry(root)
    SEMESTER_field = Entry(root)
    FORM_NO_field = Entry(root)
    CONTACT_NO_field = Entry(root)
    EMAIL_ID_field = Entry(root)
    ADDRESS_field = Entry(root)

    NAME_field.bind("<Return>",focus1)
    COURSE_field.bind("<Return>", focus2)
    SEMESTER_field.bind("<Return>", focus3)
    FORM_NO_field.bind("<Return>", focus4)
    CONTACT_NO_field.bind("<Return>", focus5)
    EMAIL_ID_field.bind("<Return>",focus6)


    NAME_field.grid(row=1,column =1,ipadx = "100")
    COURSE_field.grid(row=2, column=1, ipadx="100")             #ipadx = dividing into vertical and horizontal
    SEMESTER_field.grid(row=3, column=1, ipadx="100")
    FORM_NO_field.grid(row=4, column=1, ipadx="100")
    CONTACT_NO_field.grid(row=5, column=1, ipadx="100")
    EMAIL_ID_field.grid(row=6,column =1,ipadx= "100")
    ADDRESS_field.grid(row=7,column =1,ipadx= "100")
    excel()
    submit = Button(root,text = "SUBMIT",fg = "BLACK",bg ="RED",command = insert )
    submit.grid(row=8,column = 1)
    root.mainloop()
